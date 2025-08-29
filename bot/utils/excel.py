# -*- coding: utf-8 -*-
"""
Экспорт XLSX: листы 'data', 'summary', 'pivot_multi', 'pivot_flat'.
Устранены:
- падения от дублей MultiIndex,
- смешанные типы колонок,
- предупреждения GroupBy.apply,
- «кривая» шапка и автоширина.
"""

from __future__ import annotations

from typing import Iterable, List, Dict, Any, Tuple
from datetime import datetime
from pathlib import Path

import pandas as pd

from ..models import Contact, Agent, RepeatTouch, TalkStatus, FlyerMethod


RU_COLUMNS = {
    "agent_id": "ID агента",
    "agent_tg": "TG ID",
    "agent_username": "Логин (@)",
    "agent_name": "Имя агента",
    "full_name": "ФИО избирателя",
    "phone": "Телефон",
    "repeat_touch": "Повторность",
    "talk_status": "Статус общения",
    "flyer_method": "Способ передачи флаера",
    "flyer_number": "Номер флаера",
    "home_voting": "Голосование на дому",
    "created_at": "Создано",
}
DATA_ORDER = [
    "agent_id","agent_tg","agent_username","agent_name",
    "full_name","phone","repeat_touch","talk_status",
    "flyer_method","flyer_number","home_voting","created_at"
]


def rows_to_dataframe(rows: Iterable[Tuple[Contact, Agent]]) -> pd.DataFrame:
    """Сырые строки -> DataFrame с русскими заголовками (лист data)."""
    data: List[Dict[str, Any]] = []
    for contact, agent in rows:
        username = getattr(agent, "username", None) if agent else None
        if username and not str(username).startswith("@"):
            username = f"@{username}"
        created = contact.created_at
        if isinstance(created, datetime):
            created = created.replace(microsecond=0)
        data.append({
            "agent_id": getattr(agent, "id", None) if agent else None,
            "agent_tg": getattr(agent, "tg_user_id", None) if agent else None,
            "agent_username": username,
            "agent_name": getattr(agent, "name", None) if agent else None,
            "full_name": contact.full_name,
            "phone": contact.phone_e164,
            "repeat_touch": _map_repeat(contact.repeat_touch),
            "talk_status": _map_status(contact.talk_status),
            "flyer_method": _map_method(contact.flyer_method),
            "flyer_number": contact.flyer_number or "",
            "home_voting": "Да" if bool(contact.home_voting) else "Нет",
            "created_at": created,
        })
    df = pd.DataFrame(data)
    df = df.reindex(columns=[c for c in DATA_ORDER if c in df.columns])
    df = df.rename(columns={k: v for k, v in RU_COLUMNS.items() if k in df.columns})
    return df


def write_excel_with_pivot(df: pd.DataFrame, path: str) -> None:
    """Пишет 4 листа: data, summary, pivot_multi, pivot_flat."""
    out_path = Path(path).as_posix()
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # data
        df.to_excel(writer, index=False, sheet_name="data")
        try:
            ws_data = writer.book["data"]; ws_data.freeze_panes = "A2"; _autosize(ws_data)
        except Exception: pass

        # summary
        summary = _build_summary(df)
        summary.to_excel(writer, index=False, sheet_name="summary")
        try:
            ws_sum = writer.book["summary"]; ws_sum.freeze_panes = "A2"; _autosize(ws_sum)
        except Exception: pass

        # pivots
        p_multi, p_flat = _build_pivots(df)
        p_multi.to_excel(writer, sheet_name="pivot_multi")
        try:
            ws = writer.book["pivot_multi"]; _style_pivot(ws)
        except Exception: pass

        p_flat.to_excel(writer, index=False, sheet_name="pivot_flat")
        try:
            ws2 = writer.book["pivot_flat"]; ws2.freeze_panes = "C2"; _autosize(ws2)
        except Exception: pass


def write_admin_summary(stats: List[Dict], path: str) -> None:
    df = pd.DataFrame(stats, columns=[
        "agent_id","agent_tg","agent_username","agent_name",
        "total","consent","refusal","no_one","hand","mailbox","none","home_yes"
    ])
    df = df.rename(columns={
        "agent_id":"ID агента","agent_tg":"TG ID","agent_username":"Логин (@)","agent_name":"Имя агента",
        "total":"Всего карточек","consent":"Согласие","refusal":"Отказ","no_one":"Никого нет",
        "hand":"Флаер на руки","mailbox":"Флаер в ящик","none":"Флаер нет","home_yes":"Голосование на дому (Да)"
    })
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="summary")
        try:
            ws = w.book["summary"]; ws.freeze_panes = "A2"; _autosize(ws)
        except Exception: pass


# ---------- helpers ----------

def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Плоская сводка по агентам без GroupBy.apply предупреждений."""
    if df.empty:
        return pd.DataFrame(columns=[
            "ID агента","Логин (@)","Имя агента",
            "Всего карточек","Согласие","Отказ","Никого нет",
            "Флаер на руки","Флаер в ящик","Флаер нет","Голосование на дому (Да)"
        ])

    # Счётчики через agg по Series — без future warning
    g = df.groupby(["ID агента","Логин (@)","Имя агента"], dropna=False, sort=False)

    res = g.agg(
        **{
            "Всего карточек": ("ФИО избирателя", "size"),
            "Согласие": ("Статус общения", lambda s: (s == "Согласие").sum()),
            "Отказ": ("Статус общения", lambda s: (s == "Отказ").sum()),
            "Никого нет": ("Статус общения", lambda s: (s == "Никого нет").sum()),
            "Флаер на руки": ("Способ передачи флаера", lambda s: (s == "На руки").sum()),
            "Флаер в ящик": ("Способ передачи флаера", lambda s: (s == "В ящик").sum()),
            "Флаер нет": ("Способ передачи флаера", lambda s: (s == "Нет").sum()),
            "Голосование на дому (Да)": ("Голосование на дому", lambda s: (s == "Да").sum()),
        }
    ).reset_index()

    int_cols = ["Всего карточек","Согласие","Отказ","Никого нет","Флаер на руки","Флаер в ящик","Флаер нет","Голосование на дому (Да)"]
    for c in int_cols:
        res[c] = res[c].astype(int)

    # стабильная сортировка
    return res.sort_values(by=["ID агента","Логин (@)"], kind="mergesort")


def _build_pivots(df: pd.DataFrame):
    """Две сводные: многоуровневая и плоская. Нормализуем до 2-уровневых колонок и убираем дубли."""
    from pandas import Categorical, MultiIndex, Index

    if df.empty:
        return pd.DataFrame([["Нет данных"]]), pd.DataFrame([["Нет данных"]])

    data = df.copy()
    data["Статус общения"] = Categorical(data["Статус общения"], categories=["Никого нет","Отказ","Согласие"], ordered=True)
    data["Повторность"] = Categorical(data["Повторность"], categories=["Первичное","Повторное"], ordered=True)
    data["Способ передачи флаера"] = Categorical(data["Способ передачи флаера"], categories=["В ящик","На руки","Нет"], ordered=True)

    p_status = pd.pivot_table(
        data, index=["ID агента","Логин (@)"], columns=["Статус общения","Повторность"],
        values="ФИО избирателя", aggfunc="size", fill_value=0, dropna=False, observed=False
    )

    p_flyer = pd.pivot_table(
        data, index=["ID агента","Логин (@)"], columns=["Способ передачи флаера"],
        values="ФИО избирателя", aggfunc="size", fill_value=0, dropna=False, observed=False
    )

    # Приводим к int
    p_status = p_status.astype(int) if not p_status.empty else p_status
    p_flyer = p_flyer.astype(int) if not p_flyer.empty else p_flyer

    # Группа «Флаеры» как MultiIndex
    if not p_flyer.empty:
        p_flyer.columns = MultiIndex.from_product([["Флаеры"], p_flyer.columns])

    totals = data.groupby(["ID агента","Логин (@)"]).size()
    # totals — всегда двухуровневый
    totals.index = pd.MultiIndex.from_tuples(totals.index, names=["ID агента","Логин (@)"])
    totals.name = ("Итого", "")

    # Собираем
    parts = [x for x in [p_status, p_flyer] if x is not None and not x.empty]
    pivot = pd.concat(parts + [totals], axis=1).fillna(0)

    # Нормализуем колонки в два уровня
    if not isinstance(pivot.columns, MultiIndex):
        pivot.columns = MultiIndex.from_tuples([(str(c), "") for c in pivot.columns])
    elif pivot.columns.nlevels == 1:
        pivot.columns = MultiIndex.from_tuples([(str(c), "") for c in pivot.columns])

    # Схлопываем дубли колонок корректно (без смешения типов уровней)
    if pivot.columns.duplicated().any():
        pivot = pivot.T.groupby(level=[0, 1]).sum().T

    pivot = pivot.astype(int)

    # Красивый порядок колонок
    plan = []
    for st in ["Никого нет","Отказ","Согласие"]:
        for rt in ["Первичное","Повторное"]:
            plan.append((st, rt))
    for fm in ["В ящик","На руки","Нет"]:
        plan.append(("Флаеры", fm))
    plan.append(("Итого",""))

    ordered = [c for c in plan if c in pivot.columns]
    tail = [c for c in pivot.columns if c not in ordered]
    pivot = pivot.loc[:, ordered + tail]

    # Плоская версия
    p_flat = pivot.copy()
    p_flat.columns = [" | ".join([str(x) for x in c if str(x)]) for c in p_flat.columns]
    p_flat = p_flat.reset_index()

    return pivot, p_flat


def _autosize(ws) -> None:
    from openpyxl.styles import Alignment, Font, Border, Side
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c); cell.alignment = center; cell.font = bold; cell.border = thin
    for col in ws.columns:
        max_len = 0; col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            l = len(str(v)) if v is not None else 0
            if l > max_len: max_len = l
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)


def _style_pivot(ws) -> None:
    from openpyxl.styles import Alignment, Font, Border, Side
    max_col = ws.max_column; max_row = ws.max_row
    if max_row >= 2:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(vertical="center")
    bold = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for r in (1, 2):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c); cell.alignment = center; cell.font = bold; cell.border = thin
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c); cell.alignment = center if c > 2 else left; cell.border = thin
    for c in range(1, max_col + 1):
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            max_len = max(max_len, len(str(v)))
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 24)
    ws.freeze_panes = "C3"


def _map_repeat(v: RepeatTouch | str | None) -> str:
    s = getattr(v, "value", v) or ""
    s = getattr(v, "name", s) or s
    s = str(s).upper()
    if s in ("PRIMARY","ПЕРВИЧНОЕ"): return "Первичное"
    if s in ("SECONDARY","ПОВТОРНОЕ"): return "Повторное"
    return ""


def _map_status(v: TalkStatus | str | None) -> str:
    s = getattr(v, "value", v) or ""
    s = getattr(v, "name", s) or s
    s = str(s).upper()
    if s in ("NO_ONE","NOBODY","НИКОГО НЕТ"): return "Никого нет"
    if s in ("REFUSAL","REFUSE","ОТКАЗ"): return "Отказ"
    if s in ("CONSENT","AGREE","СОГЛАСИЕ"): return "Согласие"
    return ""


def _map_method(v: FlyerMethod | str | None) -> str:
    s = getattr(v, "value", v) or ""
    s = getattr(v, "name", s) or s
    s = str(s).upper()
    if s in ("HAND","HANDS","НА РУКИ"): return "На руки"
    if s in ("MAILBOX","BOX","В ЯЩИК"): return "В ящик"
    if s in ("NONE","НЕТ"): return "Нет"
    return ""
