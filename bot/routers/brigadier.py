from __future__ import annotations

import io
from datetime import datetime

from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from sqlalchemy import select

# pandas опционально: если нет — используем openpyxl
try:
    import pandas as pd
except Exception:
    pd = None

from ..db import async_session
from ..models import Agent
from ..repo import (
    get_or_create_agent,
    ensure_brig_tables,
    is_brigadier_allowed,
    set_brig_login,
    is_brig_logged_in,
    get_agent_by_username,
    set_brig_member,
    remove_brig_member,
    list_brigadier_member_agent_ids,
    agents_stats_for_period,
    block_member_by_username,
)
from ..keyboards import (
    kb_brig_menu, kb_brig_blacklist, kb_export_ranges,
    BTN_BRIG_MENU, BTN_BRIG_MEMBERS, BTN_BRIG_ATTACH, BTN_BRIG_DETACH,
    BTN_BRIG_BLACKLIST, BTN_BRIG_BLOCK, BTN_BRIG_UNBLOCK,
    BTN_BRIG_STATS, BTN_BRIG_EXPORT_XLSX, BTN_BRIG_LOGOUT, BTN_BRIG_HELP,
    BTN_BACK, BTN_ACCESS, BTN_BRIG_LOGIN,
    BTN_EXP_TODAY, BTN_EXP_7, BTN_EXP_30, BTN_EXP_ALL,
)

router = Router(name="brigadier")


# ===== Состояния =====
class BrigAuth(StatesGroup):
    waiting_id = State()


class BrigAttachUser(StatesGroup):
    waiting_username = State()


class BrigDetachUser(StatesGroup):
    waiting_username = State()


class BrigBlockUser(StatesGroup):
    waiting_username = State()


class BrigStats(StatesGroup):
    waiting_range = State()

async def _main_kb_for(user_id: int):
    async with async_session() as session:
        agent = await get_or_create_agent(session, user_id)
        await ensure_brig_tables(session)
        await session.commit()
        admin = bool(getattr(agent, "admin_logged_in", False))
        brig_logged = await is_brig_logged_in(session, user_id)
    return kb_main(is_admin=admin, is_brig=brig_logged)

# -------- Access --------
@router.message(F.text == BTN_ACCESS)
async def access_menu(m: Message):
    async with async_session() as session:
        await ensure_brig_tables(session)
        brig_logged = await is_brig_logged_in(session, m.from_user.id)
    await m.answer("🔑 Доступ:", reply_markup=kb_access_menu(brig_logged=brig_logged, admin_logged=False))

# -------- Login/logout --------
@router.message(F.text == BTN_BRIG_LOGIN)
async def brig_login_start(m: Message, state: FSMContext):
    await state.set_state(BrigAuth.waiting_id)
    await m.answer("🧑‍✈️ Вход бригадира.\nВведите <b>ваш ID</b> (число).")

@router.message(BrigAuth.waiting_id)
async def brig_login_finish(m: Message, state: FSMContext):
    raw = (m.text or "").strip()
    if not raw.isdigit():
        await m.answer("Нужны только цифры. Введите ваш ID.")
        return
    entered_id = int(raw)

    async with async_session() as session:
        await ensure_brig_tables(session)
        allowed = await is_brigadier_allowed(session, m.from_user.id)
        if not allowed:
            await state.clear()
            await m.answer("⛔️ Вам не назначена роль бригадира.")
            return

        agent = await get_or_create_agent(session, m.from_user.id)
        if agent.id != entered_id:
            await state.clear()
            await m.answer("❌ ID не совпадает. Проверьте у администратора.")
            return

        await set_brig_login(session, m.from_user.id, True)
        await session.commit()

    await state.clear()
    await m.answer("✅ Вход выполнен.", reply_markup=kb_brig_menu())
    await m.answer("Главное меню:", reply_markup=await _main_kb_for(m.from_user.id))

@router.message(F.text == BTN_BRIG_LOGOUT)
async def brig_logout(m: Message, state: FSMContext):
    async with async_session() as session:
        await ensure_brig_tables(session)
        await set_brig_login(session, m.from_user.id, False)
        await session.commit()
    await state.clear()
    await m.answer("🚪 Режим бригадира выключен.", reply_markup=await _main_kb_for(m.from_user.id))

@router.message(F.text == BTN_BRIG_BLACKLIST)
async def brig_blacklist_menu(m: Message):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир через «🔑 Доступ».")
            return
    await m.answer("🧱 Чёрный список — выберите действие:", reply_markup=kb_brig_blacklist())

@router.message(F.text == BTN_BRIG_MENU)
async def brig_menu(m: Message):
    async with async_session() as session:
        await ensure_brig_tables(session)
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир через «🔑 Доступ».")
            return
    await m.answer("🧑‍✈️ Меню бригадира", reply_markup=kb_brig_menu())

# -------- Members: combined list --------
@router.message(F.text == BTN_BRIG_MEMBERS)
async def brig_list_members(m: Message):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир через «🔑 Доступ».")
            return
        agent_ids = await list_brigadier_member_agent_ids(session, m.from_user.id)
        if not agent_ids:
            await m.answer("Пока нет прикреплённых участников.", reply_markup=kb_brig_menu())
            return
        rows = await session.execute(
            # Получим username/имя для списка
            __import__('sqlalchemy').select(Agent).where(Agent.id.in_(agent_ids))
        )
        agents = rows.scalars().all()
    lines = ["<b>👥 Ваши участники</b>"]
    for a in agents:
        uname = f"@{a.username}" if a.username else "(без @)"
        name = a.name or ""
        lines.append(f"• {uname} — {name} (ID {a.id})")
    lines.append("\nДействия: «Привязать», «Отвязать», «Заблокировать».")
    await m.answer("\n".join(lines), reply_markup=kb_brig_menu())

# -------- Attach by @username --------
@router.message(F.text == BTN_BRIG_ATTACH)
async def brig_attach_by_username_ask(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
    await state.set_state(BrigAttachUser.waiting_username)
    await m.answer("Введите @username участника, которого хотите <b>привязать</b> к себе.")

@router.message(BrigAttachUser.waiting_username)
async def brig_attach_by_username_save(m: Message, state: FSMContext):
    uname = (m.text or "").strip().lstrip("@")
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await state.clear()
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
        agent = await get_agent_by_username(session, uname)
        if not agent:
            await m.answer("Не нашёл такого @username. Человек должен написать боту хотя бы раз.")
            return
        await set_brig_member(session, brig_tg_id=m.from_user.id, member_tg_id=int(agent.tg_user_id))
        await session.commit()
    await state.clear()
    await m.answer(f"✅ Участник @{uname} привязан к вам.", reply_markup=kb_brig_menu())

# -------- Detach by @username --------
@router.message(F.text == BTN_BRIG_DETACH)
async def brig_detach_by_username_ask(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
    await state.set_state(BrigDetachUser.waiting_username)
    await m.answer("Введите @username участника, которого хотите <b>отвязать</b> от себя.")

@router.message(BrigDetachUser.waiting_username)
async def brig_detach_by_username_save(m: Message, state: FSMContext):
    uname = (m.text or "").strip().lstrip("@")
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await state.clear()
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
        agent = await get_agent_by_username(session, uname)
        if not agent:
            await m.answer("Не нашёл такого @username.")
            return
        await remove_brig_member(session, brig_tg_id=m.from_user.id, member_tg_id=int(agent.tg_user_id))
        await session.commit()
    await state.clear()
    await m.answer(f"🧹 Участник @{uname} отвязан.", reply_markup=kb_brig_menu())

# -------- Block by @username --------
@router.message(F.text == BTN_BRIG_BLOCK)
async def brig_block_ask(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
    await state.set_state(BrigBlockUser.waiting_username)
    await m.answer("Введите @username участника, которого хотите <b>заблокировать</b> в чате/боте.")

@router.message(BrigBlockUser.waiting_username)
async def brig_block_save(m: Message, state: FSMContext):
    uname = (m.text or "").strip().lstrip("@")
    async with async_session() as session:
        try:
            tg_id = await block_member_by_username(session, uname, blocked_by=m.from_user.id)
            await session.commit()
        except ValueError:
            await m.answer("Не нашёл такого @username. Человек должен написать боту хотя бы раз.")
            return
    await state.clear()
    await m.answer(f"🚫 @{uname} заблокирован. Доступ к опросам закрыт.", reply_markup=kb_brig_menu())

# -------- Stats for members + CSV export --------
@router.message(F.text == BTN_BRIG_STATS)
async def brig_stats_start(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
    await state.set_state(BrigStats.waiting_range)
    await m.answer("За какой период показать сводку по вашим участникам?", reply_markup=kb_export_ranges())

@router.message(F.text == BTN_BRIG_EXPORT_XLSX)
async def brig_export_start(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир через «🔑 Доступ».")
            return
    await state.set_state(BrigStats.waiting_range)
    await state.update_data(export_only=True)   # только файл, без текстовой простыни
    await m.answer("За какой период выгрузить XLSX по вашим участникам?", reply_markup=kb_export_ranges())

class BrigUnblockUser(StatesGroup):
    waiting_username = State()

@router.message(F.text == BTN_BRIG_UNBLOCK)
async def brig_unblock_ask(m: Message, state: FSMContext):
    async with async_session() as session:
        if not await is_brig_logged_in(session, m.from_user.id):
            await m.answer("⛔️ Сначала войдите как бригадир.")
            return
    await state.set_state(BrigUnblockUser.waiting_username)
    await m.answer("Введите @username участника, которого нужно ♻️ <b>разблокировать</b>.")

@router.message(BrigUnblockUser.waiting_username)
async def brig_unblock_save(m: Message, state: FSMContext):
    uname = (m.text or "").strip().lstrip("@")
    async with async_session() as session:
        try:
            await unblock_member_by_username(session, uname)
            await session.commit()
        except ValueError:
            await m.answer("Не нашёл такого @username. Человек должен один раз написать боту.")
            return
    await state.clear()
    await m.answer(f"♻️ @{uname} разблокирован. Доступ к опросам восстановлен.", reply_markup=kb_brig_menu())

@router.message(F.text == BTN_BRIG_HELP)
async def brig_help(m: Message):
    text = (
        "ℹ️ <b>Помощь бригадиру</b>\n"
        "— «Список участников» — показать привязанных к вам.\n"
        "— «Привязать/Отвязать участника» — управление вашей бригадой.\n"
        "— «Сводка по участникам» — статистика за период.\n"
        "— «Экспорт XLSX» — выгрузка данных.\n\n"
        "Если что-то не работает — нажмите «Назад» и попробуйте ещё раз."
    )
    await m.answer(text, reply_markup=kb_brig_menu())

@router.message(
    BrigStats.waiting_range,
    F.text.in_([BTN_EXP_TODAY, BTN_EXP_7, BTN_EXP_30, BTN_EXP_ALL, BTN_BACK]),
)
async def brig_stats_run(m: Message, state: FSMContext):
    if m.text == BTN_BACK:
        await state.clear()
        await m.answer("🧑‍✈️ Меню бригадира", reply_markup=kb_brig_menu())
        return

    data_state = await state.get_data()
    export_only = bool(data_state.get("export_only", False))

    # период
    days = None
    title = "за весь период"
    if m.text == BTN_EXP_TODAY:
        days, title = 1, "за 1 день"
    elif m.text == BTN_EXP_7:
        days, title = 7, "за 7 дней"
    elif m.text == BTN_EXP_30:
        days, title = 30, "за 30 дней"

    # статистика ТОЛЬКО по подопечным этого бригадира
    async with async_session() as session:
        agent_ids = await list_brigadier_member_agent_ids(session, m.from_user.id)
        all_stats = await agents_stats_for_period(session, days)
        await session.commit()

    stats = [s for s in all_stats if s.get("agent_id") in set(agent_ids)]
    if not stats:
        await state.clear()
        await m.answer("Пока нет данных по вашим участникам за выбранный период.", reply_markup=kb_brig_menu())
        return

    # Текстовая сводка (только если НЕ режим "только экспорт")
    if not export_only:
        lines = [f"<b>📈 Сводка по вашим участникам ({title})</b>"]
        for s in stats:
            uname = s.get("agent_username") or "(без @)"
            name = s.get("agent_name") or ""
            header = f"{uname} {name}".strip()
            lines.append(
                f"{header}\n"
                f"  Всего: {s.get('total',0)} | Согласие: {s.get('consent',0)} | Отказ: {s.get('refusal',0)} | Никого нет: {s.get('no_one',0)}\n"
                f"  Флаеры — На руки: {s.get('hand',0)} | В ящик: {s.get('mailbox',0)} | Нет: {s.get('none',0)} | Надомка (Да): {s.get('home_yes',0)}"
            )
        await m.answer("\n\n".join(lines))

    # ===== XLSX в памяти с русскими заголовками =====
    headers_ru = [
        "Логин (@)", "Имя",
        "Всего", "Согласие", "Отказ", "Никого нет",
        "Флаер: на руки", "Флаер: в ящик", "Флаер: не выдавали",
        "Надомное голосование (Да)",
    ]
    # порядок значений соответствует заголовкам:
    rows = []
    for s in stats:
        rows.append([
            s.get("agent_username") or "",
            s.get("agent_name") or "",
            s.get("total", 0), s.get("consent", 0), s.get("refusal", 0), s.get("no_one", 0),
            s.get("hand", 0), s.get("mailbox", 0), s.get("none", 0), s.get("home_yes", 0),
        ])

    bio = io.BytesIO()
    if pd is not None:
        # pandas + xlsxwriter → красивое форматирование
        try:
            df = pd.DataFrame(rows, columns=headers_ru)
            with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
                df.to_excel(writer, index=False, sheet_name="Сводка")
                ws = writer.sheets["Сводка"]
                wb = writer.book

                # оформление шапки
                header_fmt = wb.add_format({"bold": True, "valign": "top", "text_wrap": True})
                ws.set_row(0, None, header_fmt)

                # заморозка первой строки и автофильтр
                ws.freeze_panes(1, 0)
                ws.autofilter(0, 0, len(df), len(headers_ru) - 1)

                # примерные ширины колонок
                widths = [18, 22, 8, 10, 10, 12, 16, 16, 18, 22]
                for i, w in enumerate(widths):
                    ws.set_column(i, i, w)
        except Exception:
            # Фолбэк: openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Сводка"
            ws.append(headers_ru)
            for r in rows:
                ws.append(r)
            # жирная шапка
            for cell in ws[1]:
                cell.font = Font(bold=True)
            # заморозка и фильтр
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:J{len(rows)+1}"
            # ширины
            widths = [18, 22, 8, 10, 10, 12, 16, 16, 18, 22]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            wb.save(bio)
    else:
        # Чистый openpyxl (если pandas нет)
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Сводка"
        ws.append(headers_ru)
        for r in rows:
            ws.append(r)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{len(rows)+1}"
        widths = [18, 22, 8, 10, 10, 12, 16, 16, 18, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(bio)

    data = bio.getvalue()
    filename = f"brig_stats_{(days or 0)}d_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await m.answer_document(BufferedInputFile(data, filename=filename))

    await state.clear()
    await m.answer("🧑‍✈️ Меню бригадира", reply_markup=kb_brig_menu())
